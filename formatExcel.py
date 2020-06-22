import openpyxl 

def mergeCells(file):
    #i=0
    book = openpyxl.load_workbook(file)
    sheet = book.get_sheet_by_name('Sheet_1')
    print (sheet.max_row)
    print (sheet.max_column)
    sheet.merge_cells('B2:I2')
    sheet.merge_cells('B3:I3')
    sheet.merge_cells('B4:I4')
    sheet.merge_cells('B5:I5')
    sheet.merge_cells('B6:I6')
    sheet.merge_cells('B7:I7')
    sheet.merge_cells('B8:I8')
    sheet.merge_cells('B9:I9')
    book.save(file)
